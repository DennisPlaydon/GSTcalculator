import openpyxl, re, math
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

#Most common purchases
multiRegex = [
    r'Z\s\w+', r'Mobil\s\w+', r'Bunnings', r'Caltex\s\w+', r'Gull', r'Countdown', r'Bp\sConnect', r'Edl\s\w+', \
     r'Pak\s\w+', r'New\sWorld', r'Inex\s', r'\w+\sEnergy', r'Pallet\sPackaging', r'Vulcan', r'Woodmart',\
      r'Workstore', r'Spraywell\s\w+', r'Tga\s\w+', r'Kmart', r'Spark\sNz',r'Motor\sCycle', r'Waste\sDis\w+',\
      r'Super\sCheap', r'Placemakers', r'Botany\sHonda', r'Yamaha\sMotorcycles', r'Repco', r'Pizza', \
      r'Apco\sCoating', r'Habitat', r'Welding\sTechnology', r'Pharmacy', r'Botany\sCentral\sPost', r'Mitre\s', \
      r'Bti', r'Csp\sCoatings', r'Enco', r'Jjrichards', r'Mcwatt']
foodRegex = [r'Pak\s\w+', r'New\sWorld', r'Countdown']
fuelRegex = [ r'Z\s\w+', r'Mobil\s\w+', r'Caltex\s\w+', r'Bp\sConnect']

zeroList = ['Runchun Wang', 'Cplaydon', 'Candice Playdon', 'Dennis Playdon', 'Inland Revenue Gst']
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

total_row = 0
total_red = 0
#Loads excel file called taxprac from Desktop
wb = openpyxl.load_workbook('C:\\Users\\Dennis\\Desktop\\transactions.xlsx')
#wb = openpyxl.load_workbook('C:\\Users\\Dennis\\Desktop\\taxprac.xlsx')
#Assumes sheet containing data will be called Sheet1, could cause problems in future
sheet = wb.active
#Titles the L and M columns
sheet['L1'] = 'Payments'
sheet['M1'] = 'Receipts'
sheet.column_dimensions['E'].width = 0
sheet.column_dimensions['G'].width = 0
sheet.column_dimensions['I'].width = 0
sheet.column_dimensions['J'].width = 0
for i in range(2, sheet.max_row+1):
    try:
        #Pulls all positive money coming in and labels it as RECEIPT
        if (sheet['H'+str(i)].value) > 0:
            #sheet['K'+str(i)] = -1
            sheet['K'+str(i)] = 'RECEIPT'
        #Checks if data should is neither RECEIPT or PAYMENT e.g. transferring money to savings acc
        elif sheet['D'+str(i)].value in zeroList:
            sheet['K'+str(i)] = 0
        else:
            #Where the regex magic happens
            #Links through each element in regex **May not be most efficient method**
            for k in range(len(multiRegex)):
                haRegex = re.compile(multiRegex[k])
                #mo1 = haRegex.search(sheet['F'+str(i)].value)
                #Some data have references in column D, some in column F so this loops through both cells
                for rowOfCellObjects in sheet['D'+str(i):'F'+str(i)]:
                    for cellObj in rowOfCellObjects:
                        mo1 = haRegex.search(cellObj.value)
                        #attempts to pull data of matching regex
                        try:
                            mo1.group()
                            sheet['K'+str(i)] = 'PAYMENT'
                            '''
                            Merely continues. A scenario could happen where the regex 
                            does not match column D but a matching regex could be in column F. 
                            Don't want to prematurely delete data
                            '''
                        except:
                            continue    
 
                '''
                try:
                    mo1 = haRegex.search(sheet['D'+str(i): 'F'+str(i)])
                    print(mo1.group())
                    sheet['K'+str(i)] = 'WORKS'
                except:
                     continue
                '''
        #Decides whether to put transaction in payment column, receipt column or no column at all
        if sheet['K'+str(i)].value == 'PAYMENT':
            sheet['L'+str(i)] = sheet['H'+str(i)].value
        elif sheet['K'+str(i)].value == 'RECEIPT':
            sheet['M'+str(i)] = sheet['H'+str(i)].value
        #If the cell is blank (ie. not payment, receipt or 0) then it fills cell with red so user can manually review the transaction
        elif sheet['K'+str(i)].value == None or sheet['K'+str(i)].value == "":
            total_red += 1
            sheet['K'+str(i)].fill = redFill
            '''
                Adds a formula to column L to save user time when manually reviewing changes.
                User only needs to enter a 1 in column K for the transaction amount to appear in column L
            '''
            sheet['L'+str(i)] = '=IF(K{0}=1, H{1}, "")'.format(str(i), str(i))
        total_row += 1
        
    except Exception as exc:
        continue

#Sums the receipts and payments column at the very bottom, leaving 2 lines of space between transactions and sum
sheet['M' + str(total_row+3)] = '=SUM(M2:M{0})'.format(str(total_row+1))
sheet['L' + str(total_row+3)] = '=SUM(L2:L{0})'.format(str(total_row+1))
sheet['N' + str(total_row+3)] = '=M{0} + L{1}'.format(str(total_row+3), str(total_row+3) )
num_filled = sheet.max_row - total_red
print("Filled {0} of {1} cells, percentage accuracy: {2}%".format(num_filled, sheet.max_row, round((num_filled/sheet.max_row)*100, 2)))
#print(sheet['K131'].value)
#print(sheet['K195'].value == "")
#saves the values to a new excel file called gst_copy which is stored on the desktop
wb.save('C:\\Users\\Dennis\\Desktop\\transactions_copy.xlsx')
#wb.save('C:\\Users\\Dennis\\Desktop\\gst_copy.xlsx')
'''
TODO 

'''